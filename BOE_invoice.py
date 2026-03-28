# bala - full final corrected version (WITH PDF MATCHING LOGIC - FIXED)

import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pdfplumber
from openpyxl import Workbook

from PIL import ImageTk

# ---------------- App State ----------------
pdf_path = None
reference_pdf_path = None

TOLERANCE = 0.1  # tolerance for numeric match

# ---------------- Helpers ----------------
def safe_cell(ws, addr: str) -> str:
    try:
        val = ws[addr].value
        return str(val).strip() if val is not None else ""
    except Exception:
        return ""

def safe_cells(ws, *cells) -> str:
    values = []
    for cell in cells:
        try:
            val = ws[cell].value
            if val:
                values.append(str(val).strip())
        except Exception:
            pass
    return " ".join(values)

import pytesseract
from PIL import Image
import fitz  # PyMuPDF

pytesseract.pytesseract.tesseract_cmd = r"C:\Users\JVG\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"

def extract_full_text_from_pdf(pdf_path):
    full_text = ""

    with pdfplumber.open(pdf_path) as pdf:
        doc = fitz.open(pdf_path)  # for OCR fallback

        for i, page in enumerate(pdf.pages):
            text = page.extract_text()

            # ✅ If normal extraction fails → use OCR
            if not text or text.strip() == "":
                pix = doc[i].get_pixmap(dpi=300)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

                text = pytesseract.image_to_string(img, config='--oem 3 --psm 6')

            if text:
                full_text += " " + text

    return re.sub(r"\s+", " ", full_text.lower())

# -------- NORMALIZATION HELPERS --------
def normalize_date_digits(val: str) -> str:
    # 16/02/2026 -> 16022026
    return re.sub(r"\D", "", val)

def extract_numbers_clean(text: str):
    # remove commas first so 42,460.80 -> 42460.80
    text = text.replace(",", "")
    return re.findall(r"\d+(?:\.\d+)?", text)

# ✅ FINAL MATCH LOGIC (WORKING)
def is_matched(field, value, reference_text):
    if not value:
        return False

    # ✅ Invoice Date match (format tolerant)
    if field == "Invoice Date":
        boe_date = normalize_date_digits(value)
        ref_dates = re.findall(r"\d{2}[./-]\d{2}[./-]\d{4}", reference_text)
        for d in ref_dates:
            if normalize_date_digits(d) == boe_date:
                return True
        return False

    # ✅ Invoice Amount match (numeric tolerance)
    if field == "Invoice Amount":
        try:
            boe_amt = float(value.replace(",", ""))
        except ValueError:
            return False

        nums = extract_numbers_clean(reference_text)
        for n in nums:
            try:
                if abs(float(n) - boe_amt) <= TOLERANCE:
                    return True
            except ValueError:
                pass
        return False
    
    # ✅ NEW: Unit Price with tolerance
    if field == "Unit Price":
        try:
            boe_val = float(value.replace(",", ""))
        except ValueError:
            return False

        nums = extract_numbers_clean(reference_text)
        for n in nums:
            try:
                if abs(float(n) - boe_val) <= TOLERANCE:
                    return True
            except ValueError:
                pass
        return False
    # ✅ NEW: Quantity with tolerance
    if field == "Quantity":
        try:
            boe_val = float(value.replace(",", ""))
        except ValueError:
            return False

        nums = extract_numbers_clean(reference_text)
        for n in nums:
            try:
                if abs(float(n) - boe_val) <= TOLERANCE:
                    return True
            except ValueError:
                pass
        return False
    if field in ["Supplier Address", "Exporter Address", "Description"]:
        words = re.findall(r"\b\w+\b", value.lower())

        # remove very small/common words
        words = [w for w in words if len(w) > 2]

        match_count = 0
        for w in words:
            if w in reference_text:
                match_count += 1
            if match_count >= 8:
                return True

        return False


    # ✅ Default text match
    return value.lower() in reference_text
def color_status(entry, var):
    val = var.get()
    if val == "Matched":
        entry.config(fg="green")
    elif val == "Not Matched":
        entry.config(fg="red")


# ---------------- GUI Callbacks ----------------



def add_row(label, var, key):
    global row
    ttk.Label(container, text=label, background="white").grid(row=row, column=0, sticky="e")
    ttk.Entry(container, textvariable=var, state="readonly", width=35).grid(row=row, column=1)

    e = ttk.Entry(container, textvariable=match_status_vars[key], state="readonly", width=14)
    e.grid(row=row, column=2)

    match_status_vars[key].trace_add("write", lambda *a, e=e, v=match_status_vars[key]: color_status(e, v))
    row += 1
def upload_pdf():
    global pdf_path
    chosen = filedialog.askopenfilename(
        title="Select BOE Invoice / Invoice PDF",
        filetypes=[("PDF files", "*.pdf")]
    )
    if chosen:
        pdf_path = chosen
        pdf_label_var.set(chosen)
        status_var.set("PDF selected. Processing...")

        # ✅ AUTO RUN CONVERSION (no button needed)
        root.after(100, convert_to_excel)

def upload_reference_pdf():
    global reference_pdf_path
    chosen = filedialog.askopenfilename(
        title="Select Reference PDF for Validation",
        filetypes=[("PDF files", "*.pdf")]
    )
    if chosen:
        reference_pdf_path = chosen
        ref_pdf_label_var.set(chosen)
        status_var.set("Reference PDF selected. Click 'Validate PDF'.")

def convert_to_excel():
    if not pdf_path or not os.path.isfile(pdf_path):
        messagebox.showerror("Error", "Please select a valid BOE Invoice PDF first.")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "ExtractedText"
        row_num = 1

        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        col = 1
                        for word in line.split():
                            ws.cell(row=row_num, column=col).value = word
                            col += 1
                        row_num += 1

        invoice_no = safe_cell(ws, "C35") 
        if invoice_no == "1":

            invoice_no = safe_cell(ws, "D35")
        invoice_date = safe_cell(ws, "C2")
        gross_weight = safe_cell(ws, "M8")

        invoice_amount = safe_cell(ws, "E35")
        if invoice_amount == "INR":
            invoice_amount = safe_cell(ws, "D35")

        Supplier_Address = safe_cells(
            ws,
            "A127","B127","C127","D127","E127",
            "A128","B128","C128",
            "A129","B129","C129","D129","A130","B130","C130","D130","A164"
        )
        Exporter_Address = safe_cells(
    ws,
    "A17","B17","C17","D17",
    "A18","B18","C18","D18",
    "A19","B19","C19","D19",
    "A20","A21"
)    
        Exporter_Address_var.set(Exporter_Address)

        if not safe_cell(ws, "C141"):
            Description = safe_cells(
                ws,
                "C142","D142","E142","F142",
                "A143","B143","C143","D143"
            )
        else:
            Description = safe_cells(
                ws,
                "C141","D141","E141","F141",
                "A142","B142","C142",
                "A143","B143","C143","D143","E143","F143","G143","H143"
            )

        CTH_number = safe_cell(ws, "B141") or safe_cell(ws, "B142")
        Unit_price = safe_cell(ws, "G141") or safe_cell(ws, "G142")
        Quantity = safe_cell(ws, "H141") or safe_cell(ws, "H142")
        No_PKG = safe_cell(ws, "J195")

        Invoice_Number_var.set(invoice_no)
        invoice_date_var.set(invoice_date)
        Gross_weight_var.set(gross_weight)
        Invoice_amount_var.set(invoice_amount)
        Supplier_Address_var.set(Supplier_Address)
        CTH_number_var.set(CTH_number)
        Unit_price_var.set(Unit_price)
        Quantity_var.set(Quantity)
        No_PKG_var.set(No_PKG)

        supplier_text.config(state="normal")
        supplier_text.delete("1.0", tk.END)
        supplier_text.insert(tk.END, Supplier_Address)
        supplier_text.config(state="disabled")

        
       #✅ Exporter Address (PASTE THIS
        exporter_text.config(state="normal")
        exporter_text.delete("1.0", tk.END)
        exporter_text.insert(tk.END, Exporter_Address)
        exporter_text.config(state="disabled")


        description_text.config(state="normal")
        description_text.delete("1.0", tk.END)
        description_text.insert(tk.END, Description)
        description_text.config(state="disabled")

        status_var.set("PDF processed successfully.")
        messagebox.showinfo("Success", "Data extracted successfully.")

    except Exception as e:
        messagebox.showerror("Error", f"Failed: {e}")

def validate_with_reference_pdf():
    if not reference_pdf_path:
        messagebox.showerror("Error", "Upload Reference PDF first.")
        return

    ref_text = extract_full_text_from_pdf(reference_pdf_path)

    fields = {
        "Invoice Number": Invoice_Number_var.get(),
        "Invoice Date": invoice_date_var.get(),
        "Gross Weight": Gross_weight_var.get(),
        "Invoice Amount": Invoice_amount_var.get(),
        "Supplier Address": Supplier_Address_var.get(),
        "Exporter Address": Exporter_Address_var.get(),
        "Description": description_text.get("1.0", tk.END),
        "CTH Number": CTH_number_var.get(),
        "Unit Price": Unit_price_var.get(),
        "Quantity": Quantity_var.get(),
        "No_PKG": No_PKG_var.get(),
    }

    for k, v in fields.items():
        match_status_vars[k].set(
            "Matched" if is_matched(k, v, ref_text) else "Not Matched"
        )

    status_var.set("Validation completed.")
    messagebox.showinfo("Validation", "Reference PDF validation completed.")
def export_validation_report():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Report"

        # Headers
        ws["A1"] = "Field Name"
        ws["B1"] = "Extracted Value"
        ws["C1"] = "Match Status"

        fields = {
            "Invoice Number": Invoice_Number_var.get(),
            "Invoice Date": invoice_date_var.get(),
            "Gross Weight": Gross_weight_var.get(),
            "Invoice Amount": Invoice_amount_var.get(),
            "Supplier Address": Supplier_Address_var.get(),
            "Exporter Address": Exporter_Address_var.get(),
            "Description": description_text.get("1.0", tk.END).strip(),
            "CTH Number": CTH_number_var.get(),
            "Unit Price": Unit_price_var.get(),
            "Quantity": Quantity_var.get(),
            "No_PKG": No_PKG_var.get(),
        }

        row_num = 2
        for field, value in fields.items():
            ws.cell(row=row_num, column=1).value = field
            ws.cell(row=row_num, column=2).value = value
            ws.cell(row=row_num, column=3).value = match_status_vars[field].get()
            row_num += 1

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Validation Report"
        )

        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Success", "Validation report exported successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"Export failed: {e}")

# ---------------- GUI ----------------
root = tk.Tk()
root.geometry("1100x650")
root.resizable(False, False)


import requests
from io import BytesIO
from PIL import Image, ImageTk

def load_icon_from_url(url, size):
    response = requests.get(url)
    img = Image.open(BytesIO(response.content))
    img = img.resize(size)
    return ImageTk.PhotoImage(img)

logo_icon = load_icon_from_url(
    "https://media.designrush.com/inspiration_images/291693/conversions/ford_logo_0_c4103a3013ad-desktop.jpg",
    (80, 40)
)
header = tk.Frame(root, bg="#030303", height=70)
header.pack(fill="x")
logo = tk.Label(header, image=logo_icon, bg="#1f3a5f")
logo.pack(side="left", padx=25, pady=10)
# ---------- Background Image (SAFE GUI FIX) ----------
# ---------- Background Image using Canvas (GUI ONLY) ----------
canvas = tk.Canvas(root, width=1100, height=650, highlightthickness=0)
canvas.pack(fill="both", expand=True)

import requests
from PIL import Image, ImageTk
from io import BytesIO

try:
    url = "https://wallpapers.com/images/hd/ford-logo-1920-x-1080-wallpaper-tawqu5rsd6ne3u0p.jpg"  # ✅ direct image
    response = requests.get(url)

    bg_img = Image.open(BytesIO(response.content))
    bg_img = bg_img.resize((1100, 650))

    root.bg_photo = ImageTk.PhotoImage(bg_img)
    canvas.create_image(0, 0, image=root.bg_photo, anchor="nw")

except Exception as e:
    print("Background image skipped:", e)
root.title("BOE Invoice → Excel Converter")

container = tk.Frame(
    canvas,
    bg="white",
    bd=2,
    relief="ridge",
    padx=20,
    pady=15
)

canvas.create_window(
    550,   # center X (half of 1100)
    325,   # center Y (half of 650)
    window=container
)

pdf_label_var = tk.StringVar(value="No PDF selected")
ref_pdf_label_var = tk.StringVar(value="No reference PDF selected")
status_var = tk.StringVar(value="Please upload a BOE Invoice PDF.")


Invoice_Number_var = tk.StringVar()
invoice_date_var = tk.StringVar()
Gross_weight_var = tk.StringVar()
Invoice_amount_var = tk.StringVar()
Supplier_Address_var = tk.StringVar()
Exporter_Address_var = tk.StringVar()
CTH_number_var = tk.StringVar()
Unit_price_var = tk.StringVar()
Quantity_var = tk.StringVar()
No_PKG_var = tk.StringVar()
match_status_vars = {k: tk.StringVar() for k in [
    "Invoice Number","Invoice Date","Gross Weight","Invoice Amount",
    "Supplier Address","Exporter Address","Description","CTH Number","Unit Price","Quantity","No_PKG"
]}

tk.Label(
    container,
    text="BOE / Invoice PDF Validation System",
    font=("Segoe UI", 16, "bold"),
    bg="white",
    fg="#2c3e50"
).grid(row=0, column=0, columnspan=5, pady=10)
style = ttk.Style()
style.theme_use("clam")

style.configure(
    "Color.TButton",
    font=("Segoe UI", 10, "bold"),
    padding=8,
    foreground="white",
    background="#4CAF50"
)

style.map(
    "Color.TButton",
    background=[
        ("active", "#45a049"),
        ("pressed", "#2e7d32")
    ]
)
ttk.Button(
    container,
    text="📄 Upload BOE PDF",
    style="Color.TButton",
    command=upload_pdf
).grid(row=1, column=0, padx=6, pady=6)
#ttk.Button(container, text="⚙ Convert To Excel", style="Color.TButton",
           #command=convert_to_excel).grid(row=1, column=1, padx=6)
ttk.Button(container, text="📑 Upload Invoice PDF", style="Color.TButton",
           command=upload_reference_pdf).grid(row=1, column=2, padx=6)

ttk.Button(container, text="✅ Validate PDF", style="Color.TButton",
           command=validate_with_reference_pdf).grid(row=1, column=3, padx=6)

ttk.Button(container, text="📤 Export Report", style="Color.TButton",
           command=export_validation_report).grid(row=1, column=4, padx=6)

row = 2
def add_row(label, var, key):
    global row
    ttk.Label(container, text=label).grid(row=row, column=0, sticky="e")
    ttk.Entry(container, textvariable=var, state="readonly", width=35).grid(row=row, column=1)
    ttk.Entry(container, textvariable=match_status_vars[key], state="readonly", width=14).grid(row=row, column=2)
    row += 1

add_row("Invoice Number:", Invoice_Number_var, "Invoice Number")
add_row("Invoice Date:", invoice_date_var, "Invoice Date")
add_row("Gross Weight:", Gross_weight_var, "Gross Weight")
add_row("Invoice Amount:", Invoice_amount_var, "Invoice Amount")

ttk.Label(container, text="Consigner Address:").grid(row=row, column=0, sticky="ne")
supplier_text = tk.Text(container, width=35, height=4, state="disabled", wrap="word")
supplier_text.grid(row=row, column=1)
ttk.Entry(container, textvariable=match_status_vars["Supplier Address"], state="readonly", width=14).grid(row=row, column=2)
row += 1

ttk.Label(container, text="Consignee Address:").grid(row=row, column=0, sticky="ne")

exporter_text = tk.Text(container, width=35, height=4, state="disabled", wrap="word")
exporter_text.grid(row=row, column=1)

ttk.Entry(
    container,
    textvariable=match_status_vars["Exporter Address"],
    state="readonly",
    width=14
).grid(row=row, column=2)

row += 1



ttk.Label(container, text="Description:").grid(row=row, column=0, sticky="ne")
description_text = tk.Text(container, width=35, height=4, state="disabled", wrap="word")
description_text.grid(row=row, column=1)
ttk.Entry(container, textvariable=match_status_vars["Description"], state="readonly", width=14).grid(row=row, column=2)
row += 1

add_row("HSN No:", CTH_number_var, "CTH Number")
add_row("Unit Price INR:", Unit_price_var, "Unit Price")
add_row("Quantity:", Quantity_var, "Quantity")
add_row("No_PKG:", No_PKG_var, "No_PKG")

ttk.Label(container, textvariable=status_var, foreground="green").grid(row=row+1, column=0, columnspan=3)

root.mainloop()